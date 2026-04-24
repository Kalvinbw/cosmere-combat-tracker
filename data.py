import os
import openpyxl

EXCEL_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "Cosmere_Adversary_Stats_Combined.xlsx")
SHEET_NAME = "All Adversaries"

COLUMNS = [
    "World", "Tier", "Adversary Name", "Type",
    "Physical Defense", "Cognitive Defense", "Spiritual Defense",
    "Health", "Focus", "Investiture",
    "Physical Skills", "Cognitive Skills", "Spiritual Skills", "Invested Skills",
    "To Hit Bonus", "DPR (Fast)", "DPR (Slow)",
]

_STRING_COLS = {"World", "Adversary Name", "Type"}

_cache = {"data": None, "mtime": None}


def load_adversaries():
    mtime = os.path.getmtime(EXCEL_PATH)
    if _cache["mtime"] == mtime:
        return _cache["data"]

    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)
    ws = wb[SHEET_NAME]
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    wb.close()

    adversaries = []
    for row in rows:
        if not row or row[0] is None:
            continue
        adv = {}
        for i, col in enumerate(COLUMNS):
            val = row[i] if i < len(row) else None
            if col not in _STRING_COLS:
                val = int(val) if val is not None else 0
            else:
                val = str(val) if val is not None else ""
            adv[col] = val
        adversaries.append(adv)

    _cache["data"] = adversaries
    _cache["mtime"] = mtime
    return adversaries


def add_adversary(data):
    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME]
    row = []
    for col in COLUMNS:
        val = data.get(col, "")
        if col not in _STRING_COLS and val != "":
            try:
                val = int(val)
            except (ValueError, TypeError):
                val = 0
        row.append(val)
    ws.append(row)
    wb.save(EXCEL_PATH)
    wb.close()
    _cache["mtime"] = None  # invalidate so next load re-reads
